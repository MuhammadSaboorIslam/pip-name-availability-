import requests, json, openpyxl
from openpyxl.styles import Font, PatternFill

# Load and set Workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet["A1"] = "Occupied names"
sheet["A1"].font = Font(bold=True, italic=True)
sheet["A1"].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
sheet["B1"] = "Available Names"
sheet["B1"].font = Font(bold=True, italic=True)
sheet["B1"].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
availableIndex = 2
occupiedIndex = 2


# Load English dictionary in json format
path = 'words_dictionary.json'
with open(path, 'r') as file:
    data = json.load(file)


def check_package_exists(package_name):
    response = requests.get(f'https://pypi.org/project/{package_name}/')
    return response.status_code == 200


for item in data:
    if item == "absolutization":
        break
    if check_package_exists(item):
        sheet[f"A{availableIndex}"] = item
        print(f'Package {item} exists.')
        availableIndex += 1
    else:
        print(f'Package {item} does not exist.')
        sheet[f"B{occupiedIndex}"] = item
        occupiedIndex += 1

workbook.save("pip names.xlsx")
